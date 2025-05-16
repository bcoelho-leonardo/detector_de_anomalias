# -*- coding: utf-8 -*-
"""
Created on Thu May 15 20:42:05 2025

@author: Admin
"""

# detector_de_anomalias_streamlit.py
# -*- coding: utf-8 -*-
"""
Versão modular para uso em Streamlit:
------------------------------------
• process_file(file_like: BytesIO)  -> bytes
  - Recebe um objeto BytesIO contendo o Excel original
  - Executa toda a lógica de detecção (LOF + dados ausentes)
  - Devolve os bytes do novo arquivo *_highlighted.xlsx

"""

import os
import numpy as np
import pandas as pd
import traceback
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from sklearn.neighbors import LocalOutlierFactor

###############################################################################
# 1) Funções auxiliares (inalteradas, exceto sem prints excessivos)
###############################################################################
def dynamic_n_neighbors(row_length, frac=0.15, min_n=2, max_n=50):
    raw_nn = int(frac * row_length)
    return max(min_n, min(raw_nn, max_n))


def detect_outliers(
    df,
    date_cols,
    last_month_idx=-1,
    frac=0.15,
    min_n=2,
    max_n=50,
    fixed_cont=0.05,
    missing_threshold=0.2
):
    lof_outliers, missing_outliers = [], []
    col_name_last = date_cols[-1]

    # % de faltantes por linha
    miss_pct = df[date_cols].isnull().mean(axis=1)

    for r_idx, (_, row) in enumerate(df.iterrows()):
        y_vals = pd.to_numeric(row[date_cols], errors="coerce").values

        # 1) faltantes incomuns
        if np.isnan(y_vals[-1]) and miss_pct.iloc[r_idx] < missing_threshold:
            missing_outliers.append(r_idx)
            continue

        # 2) LOF
        if not np.isnan(y_vals[-1]):
            x_num = np.arange(len(date_cols))
            data_2d = np.column_stack([x_num, y_vals])

            valid = ~np.isnan(y_vals)
            data_2d = data_2d[valid]
            if len(data_2d) < 2:
                continue

            lof = LocalOutlierFactor(
                n_neighbors=dynamic_n_neighbors(len(data_2d), frac, min_n, max_n),
                contamination=fixed_cont,
                metric="minkowski",
                p=1,
            )
            labels = lof.fit_predict(data_2d)
            if labels[-1] == -1:  # última posição é outlier
                lof_outliers.append(r_idx)

    return lof_outliers, missing_outliers


def highlight_workbook_in_memory(
    wb,
    sheet_name,
    lof_rows,
    miss_rows,
    excel_row_offset,
    excel_col_idx=2,  # B
):
    ws = wb[sheet_name]
    yellow = PatternFill("solid", fgColor="FFFF00")
    red = PatternFill("solid", fgColor="FF0000")

    for r in lof_rows:
        ws.cell(row=r + excel_row_offset, column=excel_col_idx).fill = yellow
    for r in miss_rows:
        ws.cell(row=r + excel_row_offset, column=excel_col_idx).fill = red

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


###############################################################################
# 2) Função principal para Streamlit
###############################################################################
def process_file(file_like):
    """
    Parameters
    ----------
    file_like : BytesIO
        Conteúdo do Excel original.

    Returns
    -------
    bytes
        Bytes do arquivo *_highlighted.xlsx.
    """
    try:
        # Debug: Check input
        if not hasattr(file_like, 'read'):
            raise ValueError(f"Input não é um objeto file-like! Tipo recebido: {type(file_like)}")
        
        # Reset file pointer to beginning and check size
        file_like.seek(0)
        try:
            file_size = len(file_like.getvalue())
            print(f"Tamanho do arquivo: {file_size} bytes")
        except Exception as size_err:
            print(f"Erro ao obter tamanho: {size_err}")
        
        file_like.seek(0)  # Reset again after getting size
        
        # Check file header to diagnose format issues
        try:
            header_bytes = file_like.read(16)
            header_hex = header_bytes.hex()
            print(f"Primeiros 16 bytes: {header_hex}")
            
            # Check for Excel signatures
            is_xlsx = header_hex.startswith("504b0304")
            is_xls = header_hex.startswith("d0cf11e0")
            
            if not (is_xlsx or is_xls):
                print(f"AVISO: Assinatura de arquivo não corresponde a Excel: {header_hex}")
            
            file_like.seek(0)  # Reset pointer after header check
        except Exception as header_err:
            print(f"Erro ao verificar cabeçalho: {header_err}")
            file_like.seek(0)  # Try to reset anyway
        
        # Try to save to temporary file first
        try:
            import tempfile
            import os
            
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                file_like.seek(0)  # Ensure we're at the start
                tmp_data = file_like.read()  # Read all data
                tmp.write(tmp_data)
                temp_path = tmp.name
            
            print(f"Arquivo salvo temporariamente em: {temp_path}")
            print(f"Tamanho do arquivo temporário: {os.path.getsize(temp_path)} bytes")
            
            # -------- 1. Leitura bruta (sem header) ----------
            print("Tentando ler o arquivo Excel do caminho temporário...")
            try:
                raw_df = pd.read_excel(temp_path, sheet_name="TD Dados", header=None, engine='openpyxl')
                print(f"Excel lido com sucesso do arquivo temporário. Shape: {raw_df.shape}")
            except Exception as e:
                print(f"Erro ao ler com openpyxl do arquivo temporário: {str(e)}")
                
                # Try different Excel engines if the first one fails
                try:
                    print("Tentando com engine alternativo 'xlrd'...")
                    raw_df = pd.read_excel(temp_path, sheet_name="TD Dados", header=None, engine='xlrd')
                    print(f"Excel lido com sucesso usando 'xlrd'. Shape: {raw_df.shape}")
                except Exception as e2:
                    print(f"Falha com 'xlrd': {str(e2)}")
                    
                    # If the above fail, try checking available sheets
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(temp_path, read_only=True)
                        available_sheets = wb.sheetnames
                        print(f"Sheets disponíveis via openpyxl: {available_sheets}")
                        if "TD Dados" not in available_sheets:
                            raise ValueError(f"A planilha 'TD Dados' não foi encontrada. Sheets disponíveis: {available_sheets}")
                    except Exception as e3:
                        print(f"Falha ao verificar sheets com openpyxl: {str(e3)}")
                    
                    # Re-raise the original error with more context
                    raise ValueError(f"Não foi possível ler o arquivo Excel. Verifique se é um arquivo Excel válido (.xlsx/.xls) e contém uma aba 'TD Dados'. Erro original: {str(e)}") from e
            
            # -------- 2. Identificar cabeçalho via "ABEL" ----
            print("Procurando 'ABEL' na coluna A...")
            mask_abel = raw_df[0].astype(str).str.strip().str.upper() == "ABEL"
            if not mask_abel.any():
                print("Valores na coluna A:", raw_df[0].astype(str).str.strip().head(20).tolist())
                raise ValueError("String 'ABEL' não encontrada na coluna A.")

            start_row = mask_abel.idxmax()
            print(f"'ABEL' encontrado na linha {start_row+1}")
            
            if start_row == 0:
                raise ValueError("Não há linha de cabeçalho acima de 'ABEL'.")

            header_row = start_row - 1
            if str(raw_df.iloc[start_row - 1, 0]).strip().upper() == "UNIDADE 1":
                header_row = start_row - 2  # cabeçalho uma linha acima
                print("'UNIDADE 1' detectado. Usando header_row =", header_row)
            else:
                print("Usando header_row =", header_row)

            df = (
                raw_df.iloc[header_row:]
                .reset_index(drop=True)
            )
            df.columns = df.iloc[0]
            df = df.iloc[1:]
            df = df.set_index(df.columns[0])
            df = df.iloc[:, :-1]  # remove última coluna (total geral)
            print(f"DataFrame preparado. Shape: {df.shape}")

            # -------- 3. Filtrar colunas de data ------------
            print("Filtrando colunas de data...")
            valid_cols = [
                c for c in df.columns
                if "total geral" not in str(c).lower()
                and "total général" not in str(c).lower()
            ]
            print(f"Colunas válidas: {len(valid_cols)}")
            
            dates = pd.to_datetime(valid_cols, format="%Y-%m", errors="coerce")
            print("Datas convertidas:")
            for c, d in zip(valid_cols[:5], dates[:5]):
                print(f"  - {c} -> {d}")
            if len(valid_cols) > 5:
                print(f"  - ... mais {len(valid_cols)-5} colunas")
                
            col_sorted = [
                c for c, d in sorted(zip(valid_cols, dates), key=lambda x: x[1] or pd.Timestamp.min)
            ]
            if len(col_sorted) > 30:
                col_sorted = col_sorted[-30:]
                print(f"Limitando para as últimas 30 colunas. Última coluna: {col_sorted[-1]}")
            else:
                print(f"Usando todas as {len(col_sorted)} colunas. Última coluna: {col_sorted[-1]}")
                
            df = df[col_sorted]
            print(f"DataFrame após filtro de colunas. Shape: {df.shape}")

            # -------- 4. Detectar outliers ------------------
            print("Detectando outliers...")
            lof_rows, miss_rows = detect_outliers(
                df=df,
                date_cols=col_sorted,
                fixed_cont=0.05,
                missing_threshold=0.2,
            )
            print(f"Outliers LOF detectados: {len(lof_rows)}")
            print(f"Outliers de dados ausentes: {len(miss_rows)}")

            # -------- 5. Destacar no Excel em memória -------
            print("Carregando novamente o workbook para destacar...")
            import openpyxl
            wb = openpyxl.load_workbook(temp_path, data_only=True)
            print(f"Workbook carregado. Sheets: {wb.sheetnames}")
            
            offset = header_row + 2  # header + 1 (0-based → Excel 1-based)
            print(f"Offset para destacar: {offset}")
            
            output_bytes = highlight_workbook_in_memory(
                wb=wb,
                sheet_name="TD Dados",
                lof_rows=lof_rows,
                miss_rows=miss_rows,
                excel_row_offset=offset,
                excel_col_idx=2,
            )
            print(f"Arquivo destacado gerado com sucesso. Tamanho: {len(output_bytes)} bytes")
            
            return output_bytes
            
        finally:
            # Clean up
            if 'temp_path' in locals() and os.path.exists(temp_path):
                try:
                    os.unlink(temp_path)
                    print("Arquivo temporário removido")
                except Exception as rm_err:
                    print(f"Erro ao remover arquivo temporário: {rm_err}")
        
    except Exception as exc:
        print("Erro em process_file:", exc)
        traceback.print_exc()
        raise
                    
###############################################################################
# 3) Modo CLI opcional
###############################################################################
if __name__ == "__main__":
    path = input("Caminho do .xlsx: ").strip('"').strip("'")
    if not (os.path.isfile(path) and path.lower().endswith(".xlsx")):
        print("Arquivo inválido.")
        exit(1)
    with open(path, "rb") as f:
        out_bytes = process_file(BytesIO(f.read()))
    out_path = os.path.splitext(path)[0] + "_highlighted.xlsx"
    with open(out_path, "wb") as f:
        f.write(out_bytes)
    print("Arquivo gerado:", out_path)
