# -*- coding: utf-8 -*-
"""
Created on Fri May 16 12:22:13 2025

@author: Admin
"""
import streamlit as st
import traceback

# Global try-except block to catch any application-level errors
try:
    from io import BytesIO
    import os
    import sys
    
    # Simplify imports and check for errors
    try:
        # Add directory to Python path to ensure module can be found
        sys.path.append(os.path.dirname(os.path.abspath(__file__)))
        from detector_de_anomalias_streamlit import process_file
        st.write("✅ Importação bem-sucedida do módulo de detecção")
    except Exception as import_err:
        st.error(f"Erro ao importar o módulo: {str(import_err)}")
        st.code(traceback.format_exc())
        st.stop()

    # Configure page
    st.set_page_config(page_title="Detector de Anomalias", layout="centered")

    # Debug info at startup
    st.title("📊 Detector de Anomalias Financeiras (último mês)")

    # Show debug version & environment info
    st.caption(f"Debug Version 1.2 | Python {sys.version} | Path: {os.path.abspath(__file__)}")
    st.write(f"Diretório atual: {os.getcwd()}")
    st.write(f"Arquivos no diretório: {os.listdir()}")

    st.markdown(
    """
    Envie sua planilha **.xlsx** (aba **TD Dados**).  
    O algoritmo destacará:  
    * 🟡 valores atípicos (LOF)  
    * 🔴 ausências incomuns no último mês
    """
    )

    # File uploader
    uploaded = st.file_uploader("Escolher arquivo Excel", type=["xlsx", "xls"])

    # When a file is uploaded
    if uploaded:
        st.write(f"File '{uploaded.name}' uploaded successfully")
        st.write(f"File type: {type(uploaded)}")
        
        # Create a copy of the uploaded file data immediately
        # Modified to be more robust with session state
        try:
            # Reset session_state key to avoid stale data
            if "file_key" not in st.session_state:
                st.session_state.file_key = 0
            
            # Increment the key when a new file is uploaded
            if uploaded is not None:
                st.session_state.file_key += 1
                key = st.session_state.file_key
                
                # Store the uploaded file data with the new key
                uploaded_data = uploaded.read()
                st.session_state[f"file_data_{key}"] = uploaded_data
                st.session_state[f"file_size_{key}"] = len(uploaded_data)
                st.session_state.current_key = key
                
                st.write(f"Dados do arquivo armazenados com sucesso - key: {key}")
        except Exception as upload_err:
            st.error(f"Erro ao processar o upload: {str(upload_err)}")
            st.code(traceback.format_exc())
        
        # Create an expander for showing file info
        with st.expander("Informações do Arquivo"):
            try:
                # Get current file data using the latest key
                current_key = st.session_state.current_key
                file_data = st.session_state[f"file_data_{current_key}"]
                file_size = st.session_state[f"file_size_{current_key}"]
                
                # Create a fresh BytesIO from the saved data
                file_bytes = BytesIO(file_data)
                st.write(f"Tamanho do arquivo: {file_size} bytes")
                
                # Basic file signature check
                file_bytes.seek(0)
                file_header = file_bytes.read(8).hex() if file_size >= 8 else ""
                file_bytes.seek(0)
                
                # Check for common file signatures
                file_type = "Unknown"
                if file_header.startswith("504b0304"):
                    file_type = "ZIP file (possibly .xlsx)"
                elif file_header.startswith("d0cf11e0"):
                    file_type = "OLE Compound Document (.xls)"
                else:
                    st.warning(f"Arquivo não parece ser um Excel. Assinatura do arquivo: {file_header}")
                
                st.write(f"Tipo de arquivo detectado: {file_type}")
                
                # Try to save and reload the file to ensure it's valid
                import tempfile
                temp_path = None
                
                try:
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                        tmp.write(file_bytes.getvalue())
                        temp_path = tmp.name
                    
                    st.write(f"Arquivo salvo temporariamente em: {temp_path}")
                    
                    # Try to open with different methods
                    try:
                        st.write("Tentando abrir com openpyxl...")
                        import openpyxl
                        wb = openpyxl.load_workbook(temp_path, read_only=True)
                        st.write(f"Sheets disponíveis (openpyxl): {wb.sheetnames}")
                        if "TD Dados" in wb.sheetnames:
                            st.success("✅ Encontrada aba 'TD Dados'")
                        else:
                            st.error("❌ Aba 'TD Dados' não encontrada!")
                            st.info("Sheets disponíveis: " + ", ".join(wb.sheetnames))
                    except Exception as e1:
                        st.warning(f"Não foi possível ler com openpyxl: {str(e1)}")
                        
                        try:
                            st.write("Tentando abrir com xlrd...")
                            import xlrd
                            xls = xlrd.open_workbook(temp_path)
                            st.write(f"Sheets disponíveis (xlrd): {xls.sheet_names()}")
                            if "TD Dados" in xls.sheet_names():
                                st.success("✅ Encontrada aba 'TD Dados'")
                            else:
                                st.error("❌ Aba 'TD Dados' não encontrada!")
                        except Exception as e2:
                            st.warning(f"Não foi possível ler com xlrd: {str(e2)}")
                        
                finally:
                    # Clean up temp file
                    if temp_path and os.path.exists(temp_path):
                        os.unlink(temp_path)
                        st.write("Arquivo temporário da validação removido")
                
            except Exception as info_err:
                st.error(f"Erro ao exibir informações do arquivo: {str(info_err)}")
                st.code(traceback.format_exc())
        
        # Provide help for preparing the file
        with st.expander("Como preparar seu arquivo Excel?"):
            st.markdown("""
            ## Requisitos para o arquivo Excel:
            
            1. **Formato**: Arquivo Excel (.xlsx) válido e não corrompido
            2. **Aba**: Deve conter uma aba chamada exatamente "TD Dados"
            3. **Estrutura**:
                - Coluna A deve conter a palavra "ABEL" (identificador de início dos dados)
                - Uma linha de cabeçalho deve estar acima da linha com "ABEL"
                - Colunas com datas devem estar no formato "YYYY-MM" (exemplo: 2025-05)
            
            Se seu arquivo não atende a esses requisitos, por favor ajuste-o e tente novamente.
            """)
        
        # Process button
        if st.button("Processar"):
            st.write("Starting processing...")
            
            # Processing
            with st.spinner("Analisando..."):
                try:
                    # Use the saved file data from session state
                    try:
                        current_key = st.session_state.current_key
                        file_data = st.session_state[f"file_data_{current_key}"]
                        file_size = st.session_state[f"file_size_{current_key}"]
                        
                        if file_size <= 0:
                            st.error("Arquivo vazio. Por favor, faça upload novamente.")
                            st.stop()
                    except Exception as sess_err:
                        st.error(f"Erro ao acessar dados da sessão: {str(sess_err)}")
                        st.code(traceback.format_exc())
                        st.stop()
                    
                    st.write(f"File size for processing: {file_size} bytes")
                    
                    # Create a fresh BytesIO
                    file_bytes = BytesIO(file_data)
                    
                    # For diagnostic - check the file header
                    file_bytes.seek(0)
                    header = file_bytes.read(16).hex() if file_size >= 16 else ""
                    st.write(f"File header: {header}")
                    file_bytes.seek(0)  # Reset again
                    
                    # Save to a temporary file
                    import tempfile
                    temp_path = None
                    try:
                        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                            # Write the file content
                            tmp.write(file_bytes.getvalue())
                            temp_path = tmp.name
                        
                        st.write(f"Temporary file saved at: {temp_path}")
                        
                        # Verify the file was written correctly
                        if os.path.exists(temp_path):
                            temp_size = os.path.getsize(temp_path)
                            st.write(f"Temporary file size: {temp_size} bytes")
                            
                            if temp_size == 0:
                                st.error("Erro: O arquivo temporário está vazio!")
                                raise ValueError("O arquivo temporário está vazio. Não foi possível copiar os dados do upload.")
                            
                            # Read the file directly from disk
                            with open(temp_path, "rb") as f:
                                temp_data = f.read()
                                st.write(f"Read {len(temp_data)} bytes from temp file")
                                
                                # Create NEW BytesIO from the disk read
                                process_bytes = BytesIO(temp_data)
                                
                                # Process the file
                                resultado = process_file(process_bytes)
                        
                        # Success path
                        st.success("Pronto! Baixe o arquivo destacado:")
                        st.download_button(
                            label="⬇️ Download Excel",
                            data=resultado,
                            file_name=uploaded.name.replace(".xlsx", "_highlighted.xlsx"),
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        )
                        st.balloons()
                        
                    finally:
                        # Clean up temp file
                        if temp_path and os.path.exists(temp_path):
                            try:
                                os.unlink(temp_path)
                                st.write("Temporary file cleaned up")
                            except Exception as cleanup_err:
                                st.warning(f"Erro ao limpar arquivo temporário: {str(cleanup_err)}")
                
                except Exception as e:
                    # Error handling
                    st.error(f"Erro ao processar o arquivo: {str(e)}")
                    
                    # Show detailed error info
                    error_details = traceback.format_exc()
                    with st.expander("Detalhes do Erro"):
                        st.code(error_details)
                    
                    # Check for common errors
                    error_msg = str(e).lower()
                    if "not a zip file" in error_msg:
                        st.warning("⚠️ O arquivo não é um arquivo Excel (.xlsx) válido. Verifique se o arquivo não está corrompido.")
                        
                        # Additional debugging
                        st.write("Tentando diagnosticar o problema...")
                        try:
                            # Create a fresh BytesIO from the saved session data
                            current_key = st.session_state.current_key
                            file_data = st.session_state[f"file_data_{current_key}"]
                            fresh_bytes = BytesIO(file_data)
                            
                            st.write(f"Dados da sessão: {len(fresh_bytes.getvalue())} bytes")
                            
                            # Check file header
                            fresh_bytes.seek(0)
                            header = fresh_bytes.read(16).hex() if len(fresh_bytes.getvalue()) >= 16 else ""
                            st.write(f"Primeiros 16 bytes: {header}")
                            
                            # Try different approach
                            import pandas as pd
                            st.write("Tentando ler com diferentes engines...")
                            
                            engines = ['openpyxl', 'xlrd', 'odf', 'pyxlsb']
                            for engine in engines:
                                try:
                                    st.write(f"Tentando com engine '{engine}'...")
                                    fresh_bytes.seek(0)
                                    df = pd.read_excel(fresh_bytes, engine=engine)
                                    st.write(f"✅ Sucesso com engine '{engine}'!")
                                    break
                                except Exception as read_err:
                                    st.write(f"❌ Falha com engine '{engine}': {str(read_err)}")
                        except Exception as diag_err:
                            st.write(f"Erro no diagnóstico: {str(diag_err)}")
                        
                    elif "td dados" in error_msg:
                        st.warning("⚠️ Verifique se seu arquivo Excel contém uma aba chamada exatamente 'TD Dados'")
                    elif "abel" in error_msg:
                        st.warning("⚠️ Verifique se a coluna A contém a palavra 'ABEL'")
                    else:
                        st.warning("⚠️ Verifique o formato do arquivo de acordo com as instruções")

except Exception as main_err:
    st.error(f"Erro na aplicação: {str(main_err)}")
    st.code(traceback.format_exc())
    st.write("Por favor, tente recarregar a página ou contate o suporte.")